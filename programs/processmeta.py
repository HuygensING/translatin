from openpyxl import load_workbook

from tf.core.helpers import console, htmlEsc
from tf.core.files import readYaml
from tf.ner.helpers import toAscii

from processhelpers import METADATA_YML, METADATA_FILE


TEMPLATE = """\
<?xml version="1.0" encoding="utf-8"?>
<?xml-model
    href="https://xmlschema.huygens.knaw.nl/translatin.rng"
    type="application/xml"
    schematypens="http://relaxng.org/ns/structure/1.0"
?>
<TEI xmlns="http://www.tei-c.org/ns/1.0" xml:id="{dramaId}" xml:lang="lat">
<teiHeader>
<fileDesc>
    <titleStmt>
        <title type="main">{titleExpanded}</title>
        <title type="sub">{titleFull}</title>
        <author>{name}</author>
        <editor>{editor}</editor>
        <respStmt>
            <resp>transcription</resp>
            <name>{transcribers}</name>
        </respStmt>
    </titleStmt>
    <publicationStmt>
        <publisher>{publisher}</publisher>
        <pubPlace>{pubPlace}</pubPlace>
        <date>{pubYear}</date>
    </publicationStmt>
    <sourceDesc>
        <bibl>
            Identifier: <name>{drama}</name><lb/>
            Short title: <title>{titleShort}</title><lb/>
            Year of first edition: <date>{firstEdition}</date><lb/>
            Author information:<lb/>
            Alias: {alias}<lb/>
            Birth: {birth}, {birthPlace}<lb/>
            Death: {death}, {deathPlace}<lb/>
            Floruit: {floruit}, {activity}<lb/>
            {linkRep}<lb/>
            <emph>{source}</emph><lb/>
            {sourceLinkRep}
        </bibl>
    </sourceDesc>
</fileDesc>
<profileDesc>
    <textClass>
        <keywords>
            <term type="diction">{diction}</term>
            <term type="genre">{genre}</term>
            <term type="acts">{acts}</term>
            <term type="chorus">{chorus}</term>
            <term type="medium">{medium}</term>
        </keywords>
    </textClass>
</profileDesc>
</teiHeader>
<text>
    <body>
        {text}
        {notes}
    </body>
</text>
</TEI>
"""


class Meta:
    def __init__(self, Process):
        self.template = TEMPLATE
        self.Process = Process
        self.metaFields = readYaml(asFile=METADATA_YML)
        self.error = False
        self.dramaById = {}
        self.dramaByName = {}

    def readMetadata(self, dramaFiles):
        if self.error:
            return

        console("Collecting excel metadata ...")

        Process = self.Process
        metaFields = self.metaFields

        dramaById = self.dramaById
        dramaByName = self.dramaByName

        goodDramaFiles = {}

        default = dict(author={}, drama={})
        self.default = default

        metadata = dict(author={}, drama={})
        self.metadata = metadata
        ws = {}

        try:
            wb = load_workbook(METADATA_FILE, data_only=True)
            ws["author"] = wb["author"]
            ws["drama"] = wb["drama"]
        except Exception as e:
            Process.warn(heading=f"\t{str(e)}")
            self.error = True
            return goodDramaFiles

        for kind in ("author", "drama"):
            (headRow, *rows) = list(ws[kind].rows)

            fields = {i: head.value for (i, head) in enumerate(headRow)}
            fieldInvHead = {v: k for (k, v) in fields.items()}
            fieldInv = {}

            thisDefault = default[kind]

            for field, fieldInfo in metaFields[kind].items():
                column = fieldInfo.column

                if column not in fieldInvHead:
                    Process.warn(
                        drama=f"Meta {kind}",
                        ln=0,
                        line=column,
                        heading="Extra column configured",
                    )
                    continue

                fieldInv[column] = field
                thisDefault[field] = fieldInfo.default

                if field in {"link", "sourceLink"}:
                    thisDefault[f"{field}Rep"] = ""

            for v, k in fieldInvHead.items():
                if v not in fieldInv:
                    Process.warn(
                        drama=f"Meta {kind}",
                        ln=0,
                        line=v,
                        heading="Column not configured",
                    )
                    del fields[k]

            for r, row in enumerate(rows):
                if not any(c.value for c in row):
                    continue

                rp = r + 1

                thisMeta = {}

                for i, cell in enumerate(row):
                    fieldRep = fields.get(i, None)

                    if fieldRep is None:
                        continue

                    field = fieldInv.get(fieldRep, None)

                    if field is None:
                        Process.warn(
                            drama=f"Meta {kind}",
                            ln=rp,
                            line=fieldRep,
                            heading="Column not configured",
                        )
                        continue

                    fieldInfo = metaFields[kind][field]
                    value = cell.value or thisDefault[field]

                    if fieldInfo.tp == "str":
                        value = htmlEsc(str(value).strip()) if value else None

                    thisMeta[field] = value

                if kind == "author":
                    authorId = thisMeta["author"]

                    if authorId is None:
                        Process.warn(
                            drama="author",
                            ln=rp,
                            line=thisMeta["name"],
                            heading="no acro",
                        )
                        continue
                    elif authorId in metadata[kind]:
                        Process.warn(
                            drama="author",
                            ln=rp,
                            line=authorId,
                            heading="acro not unique",
                        )
                        continue

                    link = thisMeta.get("link", None)
                    linkRep = f"""<ref target="{link}">Link</ref>""" if link else ""
                    thisMeta["linkRep"] = linkRep

                    metadata[kind][authorId] = thisMeta

                elif kind == "drama":
                    sourceLink = thisMeta.get("sourceLink", None)
                    sourceLinkRep = (
                        f"""<ref target="{sourceLink}">Source link</ref>"""
                        if sourceLink
                        else ""
                    )
                    thisMeta["sourceLinkRep"] = sourceLinkRep

                    author = thisMeta["author"] or "Unknown"
                    title = thisMeta["titleShort"]
                    drama = toAscii(f"{author}-{title}")

                    if author != "Unknown" and author not in metadata["author"]:
                        Process.warn(
                            drama=drama,
                            ln=rp,
                            line=author,
                            heading="author not found",
                        )

                    if drama in dramaFiles:
                        realDrama = dramaFiles[drama]
                        goodDramaFiles[drama] = realDrama

                        thisMeta["drama"] = drama
                        dramaId = f"translatin{rp:>04}"
                        metadata[kind][dramaId] = thisMeta
                        dramaById[dramaId] = drama
                        dramaByName[drama] = dramaId

                    else:
                        Process.warn(drama=drama, ln=rp, heading="metadata but no data")

        for drama in sorted(dramaFiles):
            if drama not in dramaByName:
                Process.warn(drama=drama, heading="data but no metadata")

        console("Metadata collected.")
        Process.showWarnings()
        return goodDramaFiles

    def fillTemplate(self, dramaName, **data):
        if self.error:
            return None

        metaFields = self.metaFields
        template = self.template
        metadata = self.metadata
        dramaByName = self.dramaByName
        dramaDefault = self.default["drama"]
        authorDefault = self.default["author"]

        dramaId = dramaByName.get(dramaName, None)

        if dramaId is None or dramaId not in metadata["drama"]:
            thisMetadata = dramaDefault
        else:
            thisMetadata = metadata["drama"][dramaId]

        if thisMetadata.get("titleFull", None) == thisMetadata.get("titleExpanded", None):
            thisMetadata["titleFull"] = ""

        author = thisMetadata["author"]

        if author is None or author == "Unknown" or author not in metadata["author"]:
            authorMetadata = authorDefault
        else:
            authorMetadata = metadata["author"][author]

        authorMetadata = {k: v for k, v in authorMetadata.items() if k != "author"}

        corpusMetadata = {k: v.value for k, v in metaFields["corpus"].items()}

        return template.format(
            dramaId=dramaId,
            **thisMetadata,
            **corpusMetadata,
            **authorMetadata,
            **data,
        )
